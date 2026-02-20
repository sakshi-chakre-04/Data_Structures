import os
import pandas as pd
import re
from pathlib import Path

def analyze_program_content(file_path):
    """Analyze program content to determine data structure type and description"""
    try:
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read()
        
        # Determine data structure type based on content
        data_structure = "Basic"
        description = "Basic program"
        
        # Check for different data structures
        if 'struct node' in content or 'class node' in content:
            if 'prev' in content:
                data_structure = "Doubly Linked List"
                description = "Doubly linked list implementation"
            else:
                data_structure = "Singly Linked List"
                description = "Singly linked list implementation"
        elif 'Array' in content or 'arr[]' in content or 'int Arr' in content:
            data_structure = "Array"
            description = "Array operations"
        elif 'Stack' in content or 'stack' in content:
            data_structure = "Stack"
            description = "Stack implementation"
        elif 'Queue' in content or 'queue' in content:
            data_structure = "Queue"
            description = "Queue implementation"
        elif 'Tree' in content or 'tree' in content or 'Binary' in content:
            data_structure = "Tree"
            description = "Tree data structure"
        elif 'Hashtable' in content or 'HashMap' in content or 'Map' in content:
            data_structure = "Hash Table"
            description = "Hash table implementation"
        elif 'String' in content or 'string' in content or 'str' in content:
            data_structure = "String"
            description = "String manipulation"
        elif 'malloc' in content or 'free' in content or 'new ' in content:
            data_structure = "Dynamic Memory"
            description = "Dynamic memory allocation"
        elif 'file' in content or 'File' in content or 'open(' in content:
            data_structure = "File Handling"
            description = "File operations"
        elif 'printf' in content and 'scanf' in content and len(content) < 500:
            data_structure = "Basic Input/Output"
            description = "Basic input output operations"
        elif 'main()' in content and len(content) < 300:
            data_structure = "Basic Program"
            description = "Basic program structure"
        elif 'Addition' in content or 'add' in content or '+' in content:
            data_structure = "Arithmetic"
            description = "Arithmetic operations"
        elif 'for(' in content or 'while(' in content:
            data_structure = "Looping"
            description = "Looping constructs"
        elif 'if(' in content or 'else' in content:
            data_structure = "Conditional"
            description = "Conditional statements"
        
        # Look for specific function names to refine description
        if 'InsertFirst' in content or 'InsertLast' in content:
            description = "Linked list insertion operations"
        elif 'DeleteFirst' in content or 'DeleteLast' in content:
            description = "Linked list deletion operations"
        elif 'Display' in content:
            description = f"{data_structure} display operations"
        elif 'Count' in content:
            description = f"{data_structure} counting operations"
        elif 'Reverse' in content:
            description = f"{data_structure} reversal operations"
        elif 'Search' in content or 'search' in content:
            description = f"{data_structure} searching operations"
        elif 'Sort' in content or 'sort' in content:
            description = f"{data_structure} sorting operations"
        
        return data_structure, description
        
    except Exception as e:
        return "Unknown", "Unable to read file"

def main():
    # Directory containing programs
    program_dir = r"c:\Users\dell\Desktop\LB\Data_Structures"
    
    # Lists to store program information
    program_numbers = []
    program_names = []
    descriptions = []
    data_structures = []
    languages = []
    file_sizes = []
    similarities = []
    
    # Get all program files
    program_files = []
    for file in os.listdir(program_dir):
        if file.startswith('program') and (file.endswith('.c') or file.endswith('.java') or file.endswith('.js') or file.endswith('.cpp')):
            program_files.append(file)
    
    # Sort files by program number
    program_files.sort(key=lambda x: int(re.search(r'program(\d+)', x).group(1)))
    
    # Analyze each program
    prev_data_structure = None
    for file in program_files:
        file_path = os.path.join(program_dir, file)
        
        # Extract program number
        match = re.search(r'program(\d+)', file)
        if match:
            prog_num = int(match.group(1))
        else:
            prog_num = 0
        
        # Get file extension for language
        if file.endswith('.c'):
            language = 'C'
        elif file.endswith('.java'):
            language = 'Java'
        elif file.endswith('.js'):
            language = 'JavaScript'
        elif file.endswith('.cpp'):
            language = 'C++'
        else:
            language = 'Unknown'
        
        # Get file size
        file_size = os.path.getsize(file_path)
        
        # Analyze content
        data_structure, description = analyze_program_content(file_path)
        
        # Check similarity with previous program
        similarity = ""
        if prev_data_structure and prev_data_structure == data_structure:
            similarity = "Similar to previous program - same data structure"
        
        # Store information
        program_numbers.append(prog_num)
        program_names.append(file)
        descriptions.append(description)
        data_structures.append(data_structure)
        languages.append(language)
        file_sizes.append(file_size)
        similarities.append(similarity)
        
        prev_data_structure = data_structure
    
    # Create DataFrame
    df = pd.DataFrame({
        'Program Number': program_numbers,
        'Program Name': program_names,
        'Language': languages,
        'Data Structure': data_structures,
        'Description': descriptions,
        'File Size (bytes)': file_sizes,
        'Similarity Note': similarities
    })
    
    # Create color mapping for data structures
    unique_structures = df['Data Structure'].unique()
    colors = {}
    
    # Assign colors to data structures
    color_palette = [
        '#FF6B6B',  # Red
        '#4ECDC4',  # Teal
        '#45B7D1',  # Blue
        '#96CEB4',  # Green
        '#FFEAA7',  # Yellow
        '#DDA0DD',  # Plum
        '#98D8C8',  # Mint
        '#F7DC6F',  # Light Yellow
        '#BB8FCE',  # Purple
        '#85C1E2',  # Light Blue
    ]
    
    for i, structure in enumerate(unique_structures):
        colors[structure] = color_palette[i % len(color_palette)]
    
    # Create Excel file with styling
    with pd.ExcelWriter('Data_Structures_Programs.xlsx', engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Programs', index=False)
        
        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Programs']
        
        # Apply colors to data structure column
        from openpyxl.styles import PatternFill
        
        for row in range(2, len(df) + 2):  # Start from row 2 (after header)
            data_structure = df.iloc[row-2]['Data Structure']
            if data_structure in colors:
                fill = PatternFill(start_color=colors[data_structure][1:], 
                                 end_color=colors[data_structure][1:], 
                                 fill_type='solid')
                for col in range(1, len(df.columns) + 1):
                    worksheet.cell(row=row, column=col).fill = fill
        
        # Adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Create a summary sheet
    summary_df = df.groupby('Data Structure').agg({
        'Program Number': 'count',
        'Program Name': lambda x: ', '.join(x.head(3))  # Show first 3 program names
    }).reset_index()
    summary_df.columns = ['Data Structure', 'Count', 'Sample Programs']
    summary_df = summary_df.sort_values('Count', ascending=False)
    
    # Add color information to summary
    summary_df['Color'] = summary_df['Data Structure'].map(colors)
    
    with pd.ExcelWriter('Data_Structures_Programs.xlsx', engine='openpyxl', mode='a') as writer:
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        
        # Apply colors to summary sheet
        workbook = writer.book
        worksheet = writer.sheets['Summary']
        
        from openpyxl.styles import PatternFill
        
        for row in range(2, len(summary_df) + 2):
            color = summary_df.iloc[row-2]['Color']
            if color:
                fill = PatternFill(start_color=color[1:], 
                                 end_color=color[1:], 
                                 fill_type='solid')
                for col in range(1, 4):  # Color first 3 columns
                    worksheet.cell(row=row, column=col).fill = fill
        
        # Adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 60)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    print(f"Excel file 'Data_Structures_Programs.xlsx' created successfully!")
    print(f"Total programs analyzed: {len(df)}")
    print(f"Unique data structures: {len(unique_structures)}")
    print("\nData Structure Summary:")
    for structure in summary_df.itertuples():
        print(f"  {structure.Data_Structure}: {structure.Count} programs")

if __name__ == "__main__":
    main()
